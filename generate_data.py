import json, unicodedata
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as C

wb = load_workbook('/mnt/user-data/uploads/ULTIMATE_Shiny_Pokemon_Living_Dex_Guide_2_0.xlsx', data_only=True)

def norm(s): return unicodedata.normalize('NFC', str(s)).lower().replace('\u2019', "'").strip()

def colvals(sheet, spcol, codecol):
    ws = wb[sheet]; out = []
    for r in range(2, ws.max_row + 1):
        sp = ws.cell(row=r, column=spcol).value
        cd = ws.cell(row=r, column=codecol).value
        if sp is not None and cd is not None and str(sp).strip():
            out.append((str(sp).strip(), str(cd).strip()))
    return out

# ---- methods: id, name, defaultRank, ranges, plus curated metadata ----
METHODS = [
 dict(id="pla", short="PLA Outbreaks", oddsPresets=[["Mass Outbreak", 158], ["Mass Outbreak \u00b7 Dex Lv 10", 152], ["Mass Outbreak \u00b7 Dex 10 + Charm", 137], ["Massive Mass Outbreak", 316], ["Massive MO \u00b7 Dex Lv 10", 293], ["Massive MO \u00b7 Dex 10 + Charm", 241], ["Regular spawn", 4096], ["Regular spawn \u00b7 Dex 10 + Charm", 819]], name="Legends: Arceus — Mass & Massive Mass Outbreaks", game="Pokémon Legends: Arceus", console="Nintendo Switch",
      desc="Outbreaks of a single species spawn in one spot on the map. Each outbreak Pokémon gets a huge number of extra shiny rolls — +25 for a Mass Outbreak, +12 for a Massive Mass Outbreak — and you can reset the spawns by returning to the village and back. Fast, hands-on, and one of the best all-around hunts in the series.",
      oddsBase="~1/158 (Mass Outbreak) · ~1/316 (Massive Mass Outbreak)", oddsCharm="~1/137 (Mass Outbreak) · ~1/241 (Massive Mass Outbreak), with Level 10 Pokédex research",
      oddsNote="Raising a species' Pokédex research to Level 10 improves its odds even without the Shiny Charm",
      ranges=[("PLA",C("I"),C("J")),("PLA",C("K"),C("L"))], rank=1),
 dict(id="za", short="Z-A Resets", oddsPresets=[["Bench/travel reset", 4096], ["Bench/travel reset + Charm", 1024], ["Hyperspace + Sparkling Lv 3 (DLC)", 1024], ["Hyperspace + Sparkling + Charm (DLC)", 585]], name="Legends Z-A — Hyperspace, Bench & Fast-Travel Resets", game="Pokémon Legends Z-A", console="Nintendo Switch",
      desc="Wild Zone spawns reroll every time you rest at a bench or fast travel, so you can re-check dozens of Pokémon in seconds — a turbo controller lets you hunt several species at once, semi-AFK. The DLC's Hyperspace Lumiose Wild Zones add boosted odds on top, especially with a Sparkling Power Level 3 meal.",
      oddsBase="1/4096 per spawn · ~1/585 – 1/1024 in DLC Hyperspace zones with Sparkling Lv. 3", oddsCharm="1/1024 per spawn",
      oddsNote="Low odds per spawn, but the sheer speed of rerolls makes it very efficient",
      ranges=[("Z-A",C("E"),C("F"))], rank=2),
 dict(id="lg", short="LGPE Catch Combo", oddsPresets=[["Combo 31+", 341], ["Combo 31+ + Charm", 293], ["Combo 31+ + Lure + Charm", 273]], name="Let's Go Pikachu / Eevee — Catch Combos", game="Pokémon: Let's Go, Pikachu! / Let's Go, Eevee!", console="Nintendo Switch",
      desc="Catch the same species over and over to build a Catch Combo. At a combo of 31+ your shiny odds are maxed for that species, and shinies sparkle visibly in the overworld, so you can just stroll around with a Lure active and watch for the glint.",
      oddsBase="1/341 at Catch Combo 31+ with a Lure", oddsCharm="1/273 at Catch Combo 31+ with a Lure",
      oddsNote="You earn the Shiny Charm by completing the Kanto Pokédex (all 150)",
      ranges=[("LGPE",C("A"),C("B")),("LGPE",C("C"),C("D"))], rank=3),
 dict(id="sv", short="SV Sandwich", oddsPresets=[["Sparkling Lv 3", 1024], ["Sparkling Lv 3 + Charm", 683], ["Outbreak (60+ KO) + Sparkling", 819], ["Outbreak + Sparkling + Charm", 512]], name="Scarlet / Violet — Shiny Sandwich (& Mass Outbreaks)", game="Pokémon Scarlet / Violet", console="Nintendo Switch",
      desc="Make a sandwich with Sparkling Power Level 3 for your target's type (needs a Herba Mystica) and every spawn of that type gets boosted shiny odds for 30 minutes. Best combined with a Mass Outbreak of your target: knock out 60 of them first, then picnic-reset the spawns. Shinies don't sparkle in the overworld here, so watch closely for the color difference.",
      oddsBase="~1/683 – 1/1024 (Sparkling Lv. 3; better inside a cleared outbreak)", oddsCharm="~1/512 – 1/819 (Sparkling Lv. 3 inside a cleared outbreak)",
      oddsNote="Paradox Pokémon can't appear in outbreaks; isolated spawn spots make targets easier to check",
      ranges=[("SV",C("I"),C("J"))], rank=4),
 dict(id="uw", short="USUM Wormholes", oddsPresets=[["Max distance, rare ring (~36%)", 3], ["Good run (~10%)", 10], ["Minimum boost (~1%)", 100], ["Legendary soft reset", 4096], ["Legendary soft reset + Charm", 1365]], name="Ultra Sun / Ultra Moon — Ultra Wormholes", game="Pokémon Ultra Sun / Ultra Moon", console="Nintendo 3DS",
      desc="Ride Solgaleo/Lunala through Ultra Space. The rarer the wormhole and the further you travel, the higher the shiny chance of the Pokémon waiting inside — up to roughly 1 in 3. Note that these regular wormhole encounters are not soft-resettable — the shiny roll is locked when you enter, so it's a pure numbers game. Legendaries and Ultra Beasts found in wormholes are the exception: those are soft-reset hunts at standard odds instead. The Pokémon pool is limited, but the odds are some of the best in the series.",
      oddsBase="1% – 36% (up to ~1/3), scaling with wormhole type and distance", oddsCharm="Same — the Shiny Charm does not affect wormhole odds",
      oddsNote="Legendaries and Ultra Beasts are not boosted by wormhole distance — soft reset those at standard odds (1/4096, or 1/1365 with the Shiny Charm)",
      ranges=[("USUM",C("D"),C("E")),("USUM",C("F"),C("G"))], rank=5),
 dict(id="g4", short="HG/SS Odd Egg", oddsPresets=[["International Odd Egg (14%)", 7], ["Japanese Odd Egg (50%)", 2], ["Cute Charm baby (~1%)", 100]], name="Gen 4 (HeartGold / SoulSilver) — Odd Egg & Breeding Tricks", game="Pokémon HeartGold / SoulSilver", console="Nintendo DS",
      desc="Old-school breeding quirks give certain baby Pokémon dramatically boosted shiny hatch rates — see the per-species rates in your guide. The Japanese Odd Egg famously hits a 50% shiny rate; international versions sit around 14%. Slow to hatch, but the odds per egg are unmatched for these specific babies.",
      oddsBase="~1/7 (14%) international · 50% Japanese Odd Egg · ~1–3% per egg for Cute Charm-era babies", oddsCharm="Not affected — these games predate the Shiny Charm",
      oddsNote="Only a small pool of baby Pokémon qualifies",
      ranges=[("Gen4CuteCharm",C("A"),C("B")),("Gen4CuteCharm",C("E"),C("F"))], rank=6),
 dict(id="sos", short="SOS Chaining", oddsPresets=[["Chain 30+", 315], ["Chain 30+ + Charm", 273]], name="Sun / Moon & Ultra Sun / Ultra Moon — SOS Chaining", game="Pokémon Sun / Moon · Ultra Sun / Ultra Moon", console="Nintendo 3DS",
      desc="Weaken a wild Pokémon so it calls allies for help, then knock out each ally so it keeps calling. Once the chain passes 30, every new ally gets extra shiny rolls. Adrenaline Orbs and a Pokémon with False Swipe make the chain manageable. Fair warning from the guide: which species can actually be called is messy — double-check before settling in.",
      oddsBase="1/315 at chain 30+", oddsCharm="1/273 at chain 30+",
      oddsNote="The ally pool is poorly documented and complicated for some species",
      ranges=[("USUM",C("I"),C("J")),("USUM",C("K"),C("L"))], rank=7),
 dict(id="h", short="Gen 6 Hordes", oddsPresets=[["Per Pok\u00e9mon", 819], ["Per Pok\u00e9mon + Charm", 273], ["Per horde of 5", 164], ["Per horde of 5 + Charm", 55]], name="Gen 6 (X / Y & Omega Ruby / Alpha Sapphire) — Horde Encounters", game="Pokémon X / Y · Omega Ruby / Alpha Sapphire", console="Nintendo 3DS",
      desc="Use Sweet Scent (or Honey) to summon a horde of five wild Pokémon at once — five shiny checks per battle instead of one. A Pokémon with a spread move like Surf clears the non-shinies quickly. Check whether your target hordes in X/Y, ORAS, or both.",
      oddsBase="1/819 per Pokémon (≈ 1/164 per horde)", oddsCharm="1/273 per Pokémon (≈ 1/55 per horde)",
      oddsNote="",
      ranges=[("XYORAS",C("F"),C("G")),("XYORAS",C("H"),C("I"))], rank=8),
 dict(id="ss", short="Dynamax Adventures", oddsPresets=[["Per Pok\u00e9mon caught", 300], ["Per Pok\u00e9mon caught + Charm", 100]], name="Sword / Shield (Crown Tundra DLC) — Dynamax Adventures", game="Pokémon Sword / Shield + The Crown Tundra", console="Nintendo Switch",
      desc="Run through a Max Raid dungeon with rental Pokémon and catch everything you beat — up to four Pokémon per run, each with its own shiny roll you only see at the very end. The star attraction: every legendary at the end of a run is a guaranteed catch, making this the classic way to shiny hunt legendaries.",
      oddsBase="1/300 per Pokémon caught", oddsCharm="1/100 per Pokémon caught",
      oddsNote="With four catches per run, a charmed run is roughly a 1-in-25 chance of at least one shiny",
      ranges=[("SWSH",C("H"),C("I")),("SWSH",C("J"),C("K"))], rank=9),
 dict(id="cf", short="Chain Fishing", oddsPresets=[["Chain 20", 100], ["Chain 20 + Charm", 96]], name="Gen 6 (X / Y & Omega Ruby / Alpha Sapphire) — Chain Fishing", game="Pokémon X / Y · Omega Ruby / Alpha Sapphire", console="Nintendo 3DS",
      desc="Reel in fish Pokémon back-to-back without moving your character or failing a reel — each consecutive catch raises the shiny chance until it caps at a chain of 20. A lead Pokémon with Suction Cups or Sticky Hold keeps the bites coming. Relaxing, low-effort, and quick to reach great odds.",
      oddsBase="1/100 at chain 20", oddsCharm="1/96 at chain 20",
      oddsNote="",
      ranges=[("XYORAS",C("M"),C("N")),("XYORAS",C("O"),C("P"))], rank=10),
 dict(id="dpr", short="BDSP Pok\u00e9Radar", oddsPresets=[["Chain 40 (Charm has no effect)", 99]], name="Brilliant Diamond / Shining Pearl — PokéRadar", game="Pokémon Brilliant Diamond / Shining Pearl", console="Nintendo Switch",
      desc="Charge the PokéRadar in tall grass and chain the same species by always entering the patches that shake the same way, four or more squares away. At a chain of 40 your odds cap, and a sparkling grass patch means the shiny is there waiting. Breaking a chain hurts, so patience and good patch-reading are the skill here.",
      oddsBase="1/99 at chain 40", oddsCharm="Same — the Shiny Charm does not affect Radar chains",
      oddsNote="",
      ranges=[("BDSP",C("A"),C("B")),("BDSP",C("C"),C("D"))], rank=11),
 dict(id="dn", short="ORAS DexNav", oddsPresets=[["Search Level 999", 476], ["Search Level 999 + Charm", 173], ["Chain bonus peak", 56], ["Chain bonus peak + Charm", 46]], name="Omega Ruby / Alpha Sapphire — DexNav", game="Pokémon Omega Ruby / Alpha Sapphire", console="Nintendo 3DS",
      desc="Sneak up on the rustling grass shown by the DexNav to chain a species. Your Search Level with that species keeps improving the odds permanently — at Search Level 999 the odds are excellent, and chain bonuses can spike them even higher. DexNav shinies also come with egg moves, potential Hidden Abilities, and guaranteed strong IVs.",
      oddsBase="1/476 at Search Level 999 · chain bonus up to ~1/56", oddsCharm="1/173 at Search Level 999 · chain bonus up to ~1/46",
      oddsNote="Search Level progress never resets, so this hunt gets better the longer you work a species",
      ranges=[("XYORAS",C("AF"),C("AG")),("XYORAS",C("AH"),C("AI"))], rank=12),
 dict(id="fs", short="Friend Safari", oddsPresets=[["Safari encounter", 819], ["Safari encounter + Charm", 585]], name="X / Y — Friend Safari", game="Pokémon X / Y", console="Nintendo 3DS",
      desc="Post-game safaris tied to your 3DS friend codes, each containing a set trio of one type. Every encounter inside has boosted shiny odds, plus a chance at Hidden Abilities and two guaranteed perfect IVs — great for shinies you actually want to use.",
      oddsBase="1/819", oddsCharm="1/585",
      oddsNote="Which species you can access depends on your friends' safaris",
      ranges=[("XYORAS",C("S"),C("T")),("XYORAS",C("U"),C("V"))], rank=13),
 dict(id="pr", short="XY Pok\u00e9Radar", oddsPresets=[["Chain 40 (Charm has no effect)", 200]], name="X / Y — PokéRadar Chaining", game="Pokémon X / Y", console="Nintendo 3DS",
      desc="The original PokéRadar chain, refined: charge the Radar, chain identical patches of shaking grass, and hunt the sparkle at chain 40. X/Y's version has friendlier odds early in the chain than the Sinnoh original, but it's still one of the more demanding hunts to execute cleanly.",
      oddsBase="1/200 at chain 40", oddsCharm="Same — the Shiny Charm does not affect Radar chains",
      oddsNote="",
      ranges=[("XYORAS",C("X"),C("Y")),("XYORAS",C("AA"),C("AB"))], rank=14),
]

pools = {}
for m in METHODS:
    pool = {}
    for sheet, sc, cc in m["ranges"]:
        for sp, cd in colvals(sheet, sc, cc):
            pool.setdefault(norm(sp), cd)
    pools[m["id"]] = pool

# ---- sprite suffix map from Shiny tab: keyword -> natdex+formid+genderid ----
ws = wb['Shiny']
sprite = {}
for r in range(2, ws.max_row + 1):
    kw = ws.cell(row=r, column=2).value
    if not kw: continue
    a = ws.cell(row=r, column=1).value or ""
    f = ws.cell(row=r, column=6).value or ""
    h = ws.cell(row=r, column=8).value or ""
    def s(v):
        if v is None: return ""
        if isinstance(v, float) and v.is_integer(): return str(int(v))
        return str(v)
    sprite[norm(kw)] = f"{s(a)}{s(f)}{s(h)}"

# ---- master list from S-Living ----
ws = wb['S-Living']
mons = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=4).value
    if not name: continue
    name = str(name).strip()
    if name == "Gen": continue   # generation separator rows from the sheet layout
    key = norm(name)
    dex = ws.cell(row=r, column=5).value
    form = ws.cell(row=r, column=8).value
    kw = ws.cell(row=r, column=14).value
    comments = ws.cell(row=r, column=13).value
    memb = {}
    for m in METHODS:
        cd = pools[m["id"]].get(key)
        if cd: memb[m["id"]] = cd
    # Build a friendly display name for regional/alternate forms (RaichuA -> Alolan Raichu)
    display = name
    formtxt = str(form).strip() if form else ""
    REGIONS = {"Alolan":"A", "Galarian":"G", "Galairan":"G", "Hisuian":"H", "Paldean":"P",
               "Paldean Combat Breed":"P", "White Stripe":"H"}
    if formtxt in REGIONS and name.endswith(REGIONS[formtxt]):
        base = name[:-1]
        if formtxt == "White Stripe":
            display = f"White-Striped {base}"; formtxt = "Hisui"
        elif formtxt == "Paldean Combat Breed":
            display = f"Paldean {base}"; formtxt = "Combat Breed"
        else:
            adj = "Galarian" if formtxt == "Galairan" else formtxt
            display = f"{adj} {base}"
            formtxt = {"Alolan": "Alola", "Galarian": "Galar", "Galairan": "Galar", "Hisuian": "Hisui", "Paldean": "Paldea"}[formtxt]
    elif formtxt == "Original":
        formtxt = ""   # the base form needs no tag
    slug = (display + ("-" + formtxt if formtxt else "")).lower()
    slug = "".join(ch if ch.isalnum() else "-" for ch in slug)
    slug = "-".join(p for p in slug.split("-") if p)
    mon = {
        "i": len(mons),
        "k": slug,
        "n": display,
        "d": int(dex) if isinstance(dex, (int, float)) else None,
        "m": memb,
    }
    if formtxt: mon["f"] = formtxt
    if comments: mon["c"] = str(comments).strip()
    sfx = sprite.get(norm(kw)) if kw else None
    if sfx: mon["s"] = sfx
    mons.append(mon)

# ---- presets (PreferredMethods rows 21-34, cols B/C/D/E, aligned to sheet method order) ----
sheet_order = ["za","sv","ss","dpr","pla","uw","sos","lg","fs","pr","dn","h","cf","g4"]
ws = wb['PreferredMethods']
presets = {"overall": {}, "switch": {}, "3ds": {}, "beforeBank": {}}
for i, mid in enumerate(sheet_order):
    r = 21 + i
    presets["overall"][mid]    = int(ws.cell(row=r, column=2).value)
    presets["switch"][mid]     = int(ws.cell(row=r, column=3).value)
    presets["3ds"][mid]        = int(ws.cell(row=r, column=4).value)
    presets["beforeBank"][mid] = int(ws.cell(row=r, column=5).value)

data = {
    "methods": [{k: m[k] for k in ("id","short","name","game","console","desc","oddsBase","oddsCharm","oddsNote","oddsPresets","rank")} for m in METHODS],
    "presets": presets,
    "mons": mons,
}

out = "const DEX_DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n"
with open("data.js", "w", encoding="utf-8") as f:
    f.write(out)
print("mons:", len(mons), " data.js bytes:", len(out))
print("sample:", json.dumps(mons[2], ensure_ascii=False))
print("sample form:", json.dumps(next(m for m in mons if m.get('f')), ensure_ascii=False))
print("presets:", presets)
